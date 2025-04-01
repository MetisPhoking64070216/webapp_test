import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import os
import tempfile
import time

def process_excel(before_file_path, template_file_path, selected_columns, cell_positions, split_column, split_method):
    df = pd.read_excel(before_file_path, skiprows=5)

    # แปลงข้อมูลในคอลัมน์ที่มีจุดทศนิยมให้เป็นจำนวนเต็ม
    for col in df.columns:
        if df[col].dtype == 'float64':  # ตรวจสอบคอลัมน์ที่เป็นทศนิยม
            # แทนที่ค่า NaN ด้วย 0 ก่อนแปลงเป็น Int64
            df[col] = df[col].fillna(0).astype('Int64')  # แปลงเป็นจำนวนเต็ม (Int64 สามารถรองรับค่า NaN ได้)

    # ถ้าผู้ใช้เลือกให้ split ข้อมูลในคอลัมน์ที่เลือก
    if split_column and split_method:
        if split_column in df.columns:
            if split_method == "Remove Numbers":
                # แยกตัวเลขออกจากข้อความ
                df[split_column] = df[split_column].str.replace(r'^\d+\s', '', regex=True)

    try:
        wb = openpyxl.load_workbook(template_file_path)
        template_sheet = wb.active
    except Exception as e:
        st.error(f"Error opening template file: {e}")
        return None

    # ลบ "Sheet1" หากมีอยู่
    if "Sheet1" in wb.sheetnames:
        wb.remove(wb["Sheet1"])

    # เก็บรูปภาพจาก template
    template_images = []
    for image in template_sheet._images:
        temp_img_path = os.path.join(tempfile.gettempdir(), f"temp_img_{len(template_images)}.png")

        with open(temp_img_path, "wb") as img_file:
            img_file.write(image._data())  # ดึงข้อมูลรูปภาพแล้วบันทึกเป็นไฟล์

        template_images.append((temp_img_path, image.anchor))  # เก็บ path และตำแหน่ง

    for i, row in df.iterrows():
        store_code = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else "Unknown"
        sheet_name = store_code[:31]  # จำกัดชื่อชีตที่ 31 ตัวอักษร

        if sheet_name not in wb.sheetnames:
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
        else:
            new_sheet = wb[sheet_name]

        # คัดลอกข้อมูลลงในเซลล์ที่เลือก
        for col_name, cell_pos in zip(selected_columns, cell_positions):
            value = row[col_name] if col_name in df.columns else ""
            new_sheet[cell_pos] = value  

        # เพิ่มรูปภาพจาก template ลงไปในชีตใหม่
        for img_path, img_anchor in template_images:
            img = Image(img_path)  # โหลดรูปภาพใหม่จากไฟล์ที่บันทึกไว้
            new_sheet.add_image(img, img_anchor)  # วางรูปที่ตำแหน่งเดิม

    # บันทึกไฟล์ output
    output_path = os.path.join(tempfile.gettempdir(), "processed_excel.xlsx")
    try:
        wb.save(output_path)
    except Exception as e:
        st.error(f"Error saving the output file: {e}")
        return None

    return output_path

st.title("\U0001F4CA ใบปะหน้าปะล่ะ")

before_file = st.file_uploader("Upload before.xlsx", type=["xlsx"])
template_file = st.file_uploader("Upload template.xlsx", type=["xlsx"])

if before_file:
    df = pd.read_excel(before_file, skiprows=5)
    if len(df) > 1:
        column_headers = df.iloc[0].tolist()  # ใช้ค่าใน row ก่อนหน้าเป็น header
        column_options = [f"{col}" for col in df.columns]
    else:
        column_options = list(df.columns)  
else:
    column_options = []

selected_columns_display = st.multiselect("เลือกคอลัมน์ที่ต้องการดึงข้อมูล", column_options)
selected_columns = [col.split(" (")[0] for col in selected_columns_display]  # ใช้ชื่อคอลัมน์จริง

cell_positions = [st.text_input(f"ตำแหน่งเซลล์สำหรับ {col}", key=col) for col in selected_columns]

# เพิ่ม UI ให้ผู้ใช้เลือกคอลัมน์ที่ต้องการ split และเลือกวิธีการ split
split_column = st.selectbox("เลือกคอลัมน์ที่ต้องการ split", column_options)
split_method = st.selectbox("เลือกวิธีการ split", ["Remove Numbers", "Other Method"])  # คุณสามารถเพิ่มวิธีการอื่น ๆ ได้ในอนาคต

if before_file and selected_columns:
    st.subheader("🔍 Preview: ข้อมูลก่อนหน้า")
    if len(df) > 1:
        df_prev = df[selected_columns].shift(1)  # เลื่อนขึ้น 1 แถวเพื่อดูค่าก่อนหน้า
        df_prev.dropna(inplace=True)  # ลบค่า NaN ที่เกิดจากการ shift
        st.dataframe(df_prev.style.set_properties(**{"background-color": "#f0f0f0", "color": "black"}))
    else:
        st.info("ไฟล์มีข้อมูลไม่พอสำหรับแสดงค่าก่อนหน้า")

if st.button("Generate Excel File"):
    if before_file and template_file and selected_columns and cell_positions:
        temp_dir = tempfile.gettempdir()
        before_file_path = os.path.join(temp_dir, "before.xlsx")
        template_file_path = os.path.join(temp_dir, "template.xlsx")
        
        with open(before_file_path, "wb") as f:
            f.write(before_file.getbuffer())
        with open(template_file_path, "wb") as f:
            f.write(template_file.getbuffer())
        
        time.sleep(1)
        
        output_file = process_excel(before_file_path, template_file_path, selected_columns, cell_positions, split_column, split_method)
        if output_file:
            with open(output_file, "rb") as file:
                st.download_button("Download Processed Excel", file, file_name="processed_excel.xlsx")
    else:
        st.error("กรุณาอัปโหลดไฟล์ เลือกคอลัมน์ และกรอกตำแหน่งเซลล์!")
